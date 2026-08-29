# backend/domain/admin/service.py — 认证 / 配置中心 / 审计
"""admin 域服务层。

事务纪律：Service 层统一 commit/rollback；审计日志与业务变更同事务写入。
配置缓存：进程内 TTL 缓存（60s），更新同键即失效——单机部署（ADR-005）足够。
"""

from __future__ import annotations

import json
from datetime import datetime

from sqlalchemy.orm import Session

from backend.common.config_service import (  # noqa: F401 — 兼容旧引用
    ConfigService,
    invalidate_config_cache,
)
from backend.common.exceptions import (
    ConflictError,
    ForbiddenError,
    NotFoundError,
    UnauthorizedError,
    ValidationError,
)
from backend.common.security import create_admin_token, hash_password, verify_password
from backend.domain.admin.models import AdminUser, AuditLog, SystemConfig
from backend.domain.admin.repository import (
    AdminUserRepository,
    AuditLogRepository,
    SystemConfigRepository,
)
from backend.domain.catalog.audit_events import publish_audit

# ---------- 权限目录（声明式 RBAC 的单一事实源） ----------
# superadmin = 全量；staff = 日常运营（借还/图书/活动/会员办理/放行留痕），不含资金审核与系统管理
STAFF_PERMISSIONS = [
    "config.view",
    "book.manage",
    "borrow.operate",
    "member.manage",
    "activity.manage",
    "quiz.manage",
    "audio.manage",
    "dashboard.view",
]
SUPER_ADMIN_PERMISSIONS = ["*"]


def permissions_for_role(role: str) -> list[str]:
    if role == AdminUser.ROLE_SUPER_ADMIN:
        return SUPER_ADMIN_PERMISSIONS
    return STAFF_PERMISSIONS


def role_has_permission(role: str, perm_code: str) -> bool:
    perms = permissions_for_role(role)
    return "*" in perms or perm_code in perms


class AuthService:
    def __init__(self, db: Session):
        self.db = db
        self.user_repo = AdminUserRepository(db)
        self.audit_repo = AuditLogRepository(db)

    def login(self, username: str, password: str) -> tuple[str, AdminUser]:
        user = self.user_repo.get_by_username(username)
        if not user or not verify_password(password, user.password_hash):
            raise UnauthorizedError("用户名或密码错误")
        if user.status != AdminUser.STATUS_ACTIVE:
            raise ForbiddenError("账号已禁用，请联系超级管理员")

        token = create_admin_token(user.id, user.role, user.token_generation)
        user.last_login_at = datetime.now()
        self.user_repo.update(user)
        self.audit_repo.create(
            AuditLog(
                actor_id=user.id,
                actor_name=user.display_name or user.username,
                action=AuditLog.ACTION_LOGIN,
                target_type="admin_user",
                target_id=str(user.id),
                detail=json.dumps(
                    {"username": user.username, "role": user.role}, ensure_ascii=False
                ),
                reason="登录",
            )
        )
        self.db.commit()
        return token, user


class AuditService:
    def __init__(self, db: Session):
        self.db = db
        self.audit_repo = AuditLogRepository(db)

    def list_logs(
        self,
        page: int = 1,
        page_size: int = 20,
        actor_id: int | None = None,
        action: str | None = None,
    ) -> tuple[list[AuditLog], int]:
        return self.audit_repo.list_with_filters(page, page_size, actor_id, action)


class DashboardService:
    """仪表盘运行数据（WM1：全部真实可查；业务指标随模块交付扩展）。"""

    def __init__(self, db: Session):
        self.db = db
        self.user_repo = AdminUserRepository(db)
        self.config_repo = SystemConfigRepository(db)
        self.audit_repo = AuditLogRepository(db)

    def overview(self) -> dict:
        from datetime import datetime, timedelta

        from sqlalchemy import func

        today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        admin_count = (
            self.db.query(func.count(AdminUser.id))
            .filter(AdminUser.is_deleted == 0, AdminUser.status == AdminUser.STATUS_ACTIVE)
            .scalar()
            or 0
        )
        today_logins = (
            self.db.query(func.count(AuditLog.id))
            .filter(
                AuditLog.action == AuditLog.ACTION_LOGIN,
                AuditLog.created_at >= today_start,
            )
            .scalar()
            or 0
        )
        config_count = (
            self.db.query(func.count(SystemConfig.id)).filter(SystemConfig.is_deleted == 0).scalar()
            or 0
        )

        # ---- 经营格子（C21：只读统计，口径与 circulation/identity/activity 域一致） ----
        from backend.domain.activity.models import ActivityEnrollment
        from backend.domain.catalog.models import BookCopy
        from backend.domain.circulation.models import BorrowRecord
        from backend.domain.identity.models import Child

        copy_total = (
            self.db.query(func.count(BookCopy.id)).filter(BookCopy.is_deleted == 0).scalar() or 0
        )
        copy_available = (
            self.db.query(func.count(BookCopy.id))
            .filter(BookCopy.is_deleted == 0, BookCopy.status == BookCopy.STATUS_AVAILABLE)
            .scalar()
            or 0
        )
        copy_borrowed = (
            self.db.query(func.count(BookCopy.id))
            .filter(BookCopy.is_deleted == 0, BookCopy.status == BookCopy.STATUS_BORROWED)
            .scalar()
            or 0
        )
        today_borrowed = (
            self.db.query(func.count(BorrowRecord.id))
            .filter(BorrowRecord.is_deleted == 0, BorrowRecord.borrowed_at >= today_start)
            .scalar()
            or 0
        )
        today_returned = (
            self.db.query(func.count(BorrowRecord.id))
            .filter(
                BorrowRecord.is_deleted == 0,
                BorrowRecord.returned_at.isnot(None),
                BorrowRecord.returned_at >= today_start,
            )
            .scalar()
            or 0
        )
        # 逾期口径与 circulation.overdue_list 保持一致（active/overdue + due_at < now）
        overdue_active = (
            self.db.query(func.count(BorrowRecord.id))
            .filter(
                BorrowRecord.is_deleted == 0,
                BorrowRecord.status.in_([BorrowRecord.STATUS_ACTIVE, BorrowRecord.STATUS_OVERDUE]),
                BorrowRecord.due_at < datetime.now(),
            )
            .scalar()
            or 0
        )
        member_total = (
            self.db.query(func.count(Child.id))
            .filter(
                Child.is_deleted == 0,
                Child.member_status.in_(
                    [
                        Child.MEMBER_OBSERVATION,
                        Child.MEMBER_PENDING_EVALUATION,
                        Child.MEMBER_FORMAL,
                        Child.MEMBER_EXPIRED,
                    ]
                ),
            )
            .scalar()
            or 0
        )
        week_start = today_start - timedelta(days=datetime.now().weekday())
        member_new_week = (
            self.db.query(func.count(Child.id))
            .filter(Child.is_deleted == 0, Child.create_time >= week_start)
            .scalar()
            or 0
        )
        activity_enroll_recent = (
            self.db.query(func.count(ActivityEnrollment.id))
            .filter(
                ActivityEnrollment.is_deleted == 0,
                ActivityEnrollment.created_at >= datetime.now() - timedelta(days=7),
            )
            .scalar()
            or 0
        )

        # ---- WM11 看板补全（D5/FEAT-069：维护/遗失/续费率/退会率/测验通过率/里程碑） ----
        from backend.domain.growth.models import MilestoneAward, QuizAttempt

        copy_maintenance = (
            self.db.query(func.count(BookCopy.id))
            .filter(BookCopy.is_deleted == 0, BookCopy.status == BookCopy.STATUS_MAINTENANCE)
            .scalar()
            or 0
        )
        copy_lost = (
            self.db.query(func.count(BookCopy.id))
            .filter(BookCopy.is_deleted == 0, BookCopy.status == BookCopy.STATUS_LOST)
            .scalar()
            or 0
        )
        returned_total = (
            self.db.query(func.count(BorrowRecord.id))
            .filter(
                BorrowRecord.is_deleted == 0, BorrowRecord.status == BorrowRecord.STATUS_RETURNED
            )
            .scalar()
            or 0
        )
        renewed_total = (
            self.db.query(func.count(BorrowRecord.id))
            .filter(
                BorrowRecord.is_deleted == 0,
                BorrowRecord.renew_used > 0,
                BorrowRecord.status == BorrowRecord.STATUS_RETURNED,
            )
            .scalar()
            or 0
        )
        renew_rate = round(renewed_total * 100 / returned_total, 1) if returned_total else 0.0
        withdrawn_total = (
            self.db.query(func.count(Child.id))
            .filter(Child.is_deleted == 0, Child.member_status == Child.MEMBER_WITHDRAWN)
            .scalar()
            or 0
        )
        withdrawal_rate = (
            round(withdrawn_total * 100 / (member_total + withdrawn_total), 1)
            if (member_total + withdrawn_total)
            else 0.0
        )
        quiz_total = (
            self.db.query(func.count(QuizAttempt.id)).filter(QuizAttempt.is_deleted == 0).scalar()
            or 0
        )
        quiz_passed = (
            self.db.query(func.count(QuizAttempt.id))
            .filter(QuizAttempt.is_deleted == 0, QuizAttempt.passed == 1)
            .scalar()
            or 0
        )
        quiz_pass_rate = round(quiz_passed * 100 / quiz_total, 1) if quiz_total else 0.0
        milestone_count = (
            self.db.query(func.count(MilestoneAward.id))
            .filter(MilestoneAward.is_deleted == 0)
            .scalar()
            or 0
        )
        pending_evaluation_count = (
            self.db.query(func.count(Child.id))
            .filter(
                Child.is_deleted == 0,
                Child.member_status == Child.MEMBER_PENDING_EVALUATION,
            )
            .scalar()
            or 0
        )

        recent_changes = (
            self.db.query(AuditLog)
            .filter(
                AuditLog.action == AuditLog.ACTION_CONFIG_UPDATE,
                AuditLog.is_deleted == 0,
            )
            .order_by(AuditLog.created_at.desc())
            .limit(5)
            .all()
        )

        def _fmt(entry: AuditLog) -> dict:
            config = self.config_repo.get_by_key(entry.target_id)
            name = config.display_name or config.description if config else entry.target_id
            try:
                detail = json.loads(entry.detail or "{}")
                change = f"{detail.get('old', '?')} → {detail.get('new', '?')}"
            except (ValueError, TypeError):
                change = entry.detail or ""
            return {
                "config_name": name or entry.target_id,
                "change": change,
                "actor_name": entry.actor_name,
                "created_at": entry.created_at.strftime("%m-%d %H:%M") if entry.created_at else "",
            }

        return {
            "admin_count": admin_count,
            "today_logins": today_logins,
            "config_count": config_count,
            "recent_config_changes": [_fmt(e) for e in recent_changes],
            "copy_total": copy_total,
            "copy_available": copy_available,
            "copy_borrowed": copy_borrowed,
            "today_borrowed": today_borrowed,
            "today_returned": today_returned,
            "overdue_active": overdue_active,
            "member_total": member_total,
            "member_new_week": member_new_week,
            "activity_enroll_recent": activity_enroll_recent,
            "copy_maintenance": copy_maintenance,
            "copy_lost": copy_lost,
            "renew_rate": renew_rate,
            "withdrawal_rate": withdrawal_rate,
            "quiz_pass_rate": quiz_pass_rate,
            "milestone_count": milestone_count,
            "pending_evaluation_count": pending_evaluation_count,
        }


class NotifyAdminService:
    """WM11 通知记录中心（管理端）。"""

    def __init__(self, db: Session):
        self.db = db

    def list_notifications(
        self,
        page: int,
        page_size: int,
        category: str | None = None,
        scene: str | None = None,
        parent_name: str | None = None,
        unread: bool | None = None,
        read: bool | None = None,
    ) -> tuple[list[dict], int, int, int]:
        """通知记录列表。返回 (items, total, unread_count, all_count)。

        - unread_count：当前筛选（category/scene/parent_name）下的未读数，Tab 计数口径（不动）；
        - all_count：当前筛选下、不含已读过滤的总数（Tab「全部（N）」计数口径，C41）；
        - unread=True 时 total 下沉为 SQL 未读数（分页页数正确，审查必修 bug）；
        - read=True 时 total 下沉为 SQL 已读数（Tab「已读」，C42）；unread/read 同传时 unread 优先；
        - unread/read 过滤是 SQL 条件而非页内 Python 过滤。
        """
        from backend.common.notification_models import Notification
        from backend.domain.identity.models import Parent

        q = self.db.query(Notification, Parent).join(
            Parent, Notification.parent_id == Parent.id, isouter=True
        )
        q = q.filter(Notification.is_deleted == 0)
        if category:
            q = q.filter(Notification.category == category)
        if scene:
            q = q.filter(Notification.scene == scene)
        if parent_name:
            q = q.filter(Parent.name.like(f"%{parent_name}%"))
        all_count = q.count()
        unread_count = q.filter(Notification.read_at.is_(None)).count()
        if unread:
            q = q.filter(Notification.read_at.is_(None))
        elif read:
            q = q.filter(Notification.read_at.is_not(None))
        total = q.count()
        rows = (
            q.order_by(Notification.id.desc()).offset((page - 1) * page_size).limit(page_size).all()
        )
        items = [
            {
                "id": n.id,
                "parent_name": p.name if p else f"#{n.parent_id}",
                "parent_id": n.parent_id,
                "child_id": n.child_id,
                "scene": n.scene,
                "category": n.category,
                "title": n.title,
                "content": n.content,
                "ref_type": n.ref_type,
                "ref_id": n.ref_id,
                "wechat_status": n.wechat_status,
                "wechat_error": n.wechat_error,
                "read": n.is_read,
                "created_at": n.create_time.strftime("%Y-%m-%d %H:%M") if n.create_time else "",
            }
            for n, p in rows
        ]
        return items, total, unread_count, all_count

    def toggle_read(self, admin, notification_id: int, read: bool, reason: str = "") -> dict:
        """管理端代家长标记已读/未读（运营介入，审计留痕——Q2 裁决）。"""
        from datetime import datetime

        from backend.common.notification_models import Notification

        n = (
            self.db.query(Notification)
            .filter(Notification.id == notification_id, Notification.is_deleted == 0)
            .first()
        )
        if not n:
            raise NotFoundError("通知不存在")
        n.read_at = datetime.now() if read else None
        self.db.flush()
        from backend.domain.catalog.audit_events import publish_audit

        publish_audit(
            self.db,
            admin=admin,
            action="notification.toggle_read",
            target_type="notification",
            target_id=str(notification_id),
            detail={"read": read, "parent_id": n.parent_id, "scene": n.scene},
            reason=reason or ("管理端标记已读" if read else "管理端标记未读"),
        )
        self.db.commit()
        # 全局口径计数随响应返回（F1b/C37）：前端 Tab 计数以服务端为准，不再本地推算
        from backend.common.notification_models import Notification as NModel

        unread_count = (
            self.db.query(NModel).filter(NModel.is_deleted == 0, NModel.read_at.is_(None)).count()
        )
        total = self.db.query(NModel).filter(NModel.is_deleted == 0).count()
        return {
            "id": notification_id,
            "read": bool(read),
            "unread_count": unread_count,
            "total": total,
        }

    def export_excel(self) -> bytes:
        """通知记录导出 Excel（与审计导出同用 openpyxl）。"""
        from io import BytesIO

        from openpyxl import Workbook
        from openpyxl.styles import Font

        rows, _, _, _ = self.list_notifications(1, 10000)
        wb = Workbook()
        ws = wb.active
        ws.title = "通知记录"
        headers = ["ID", "家长", "分类", "场景", "标题", "内容", "微信状态", "已读", "时间"]
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True)
        for r in rows:
            ws.append(
                [
                    r["id"],
                    r["parent_name"],
                    r["category"],
                    r["scene"],
                    r["title"],
                    r["content"],
                    r["wechat_status"],
                    "是" if r["read"] else "否",
                    r["created_at"],
                ]
            )
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()


class TaskAdminService:
    """WM11 定时任务看板（任务清单/运行记录/手动触发）。"""

    def __init__(self, db: Session):
        self.db = db

    def specs(self) -> list[dict]:
        from backend.common.notification_models import TaskRunLog
        from backend.tasks.registry import list_task_specs

        specs = list_task_specs()
        # 每个任务的最新一条运行记录（Q2 裁决：last_run join；空值前端显示"从未运行"）
        by_task: dict[str, dict] = {}
        rows = (
            self.db.query(TaskRunLog)
            .filter(TaskRunLog.is_deleted == 0)
            .order_by(TaskRunLog.id.desc())
            .limit(500)
            .all()
        )
        for r in rows:
            by_task.setdefault(
                r.task_name,
                {
                    "status": r.status,
                    "processed": r.processed,
                    "error": r.error,
                    "started_at": r.started_at.strftime("%Y-%m-%d %H:%M:%S")
                    if r.started_at
                    else "",
                },
            )
        for s in specs:
            last = by_task.get(s["name"])
            s["last_run"] = last or None
        return specs

    def recent_runs(self, limit: int = 20) -> list[dict]:
        from backend.common.notification_models import TaskRunLog

        rows = (
            self.db.query(TaskRunLog)
            .filter(TaskRunLog.is_deleted == 0)
            .order_by(TaskRunLog.id.desc())
            .limit(limit)
            .all()
        )
        return [
            {
                "task_name": r.task_name,
                "status": r.status,
                "processed": r.processed,
                "error": r.error,
                "started_at": r.started_at.strftime("%Y-%m-%d %H:%M:%S") if r.started_at else "",
                "finished_at": r.finished_at.strftime("%Y-%m-%d %H:%M:%S") if r.finished_at else "",
            }
            for r in rows
        ]

    def run(self, task_name: str, manual: bool = False, admin=None) -> dict:
        from backend.tasks.registry import run_task

        return run_task(task_name, manual=manual, admin=admin)


class AuditExportService:
    """审计日志 Excel 导出（C18：FEAT-005「可查询导出」）。"""

    def __init__(self, db: Session):
        self.db = db

    def export_excel(self) -> bytes:
        from io import BytesIO

        from openpyxl import Workbook
        from openpyxl.styles import Font

        rows = (
            self.db.query(AuditLog)
            .filter(AuditLog.is_deleted == 0)
            .order_by(AuditLog.id.desc())
            .limit(10000)
            .all()
        )
        wb = Workbook()
        ws = wb.active
        ws.title = "操作审计"
        headers = ["ID", "操作人", "动作", "对象类型", "对象ID", "详情", "原因", "时间"]
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True)
        for r in rows:
            ws.append(
                [
                    r.id,
                    r.actor_name,
                    r.action,
                    r.target_type,
                    r.target_id,
                    r.detail or "",
                    r.reason,
                    r.created_at.strftime("%Y-%m-%d %H:%M:%S") if r.created_at else "",
                ]
            )
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()


class DashboardExportService:
    """数据看板 Excel 导出（docs/04 WM11 步骤 7：导出任意报表 Excel）。"""

    def __init__(self, db: Session):
        self.db = db

    def export_excel(self) -> bytes:
        from io import BytesIO

        from openpyxl import Workbook
        from openpyxl.styles import Font

        data = DashboardService(self.db).overview()
        wb = Workbook()
        ws = wb.active
        ws.title = "数据看板"
        ws.append(["指标", "数值"])
        for c in ws[1]:
            c.font = Font(bold=True)
        label_map = {
            "admin_count": "后台账号",
            "today_logins": "今日登录",
            "config_count": "业务配置项",
            "copy_total": "总藏书量",
            "copy_available": "在馆",
            "copy_borrowed": "借出",
            "copy_maintenance": "维护",
            "copy_lost": "遗失",
            "today_borrowed": "今日借出",
            "today_returned": "今日归还",
            "overdue_active": "当前逾期",
            "member_total": "会员总数",
            "member_new_week": "本周新增会员",
            "pending_evaluation_count": "待评估人数",
            "activity_enroll_recent": "近7天活动报名",
            "renew_rate": "续费率(%)",
            "withdrawal_rate": "退会率(%)",
            "quiz_pass_rate": "测验通过率(%)",
            "milestone_count": "里程碑达成人数",
        }
        for key, label in label_map.items():
            ws.append([label, data.get(key, 0)])
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()


class StaffService:
    """员工账号管理（WM1 超管职责：创建/禁用/改角色/重置密码）。"""

    def __init__(self, db: Session):
        self.db = db
        self.user_repo = AdminUserRepository(db)

    def list(self) -> list[AdminUser]:
        return (
            self.db.query(AdminUser)
            .filter(AdminUser.is_deleted == 0)
            .order_by(AdminUser.id.asc())
            .all()
        )

    def create(
        self, admin: AdminUser, username: str, password: str, display_name: str, role: str
    ) -> AdminUser:
        if self.user_repo.get_by_username(username):
            raise ConflictError("用户名已存在")
        if role not in (AdminUser.ROLE_SUPER_ADMIN, AdminUser.ROLE_STAFF):
            raise ValidationError("角色必须是 superadmin 或 staff")
        user = AdminUser(
            username=username,
            password_hash=hash_password(password),
            display_name=display_name,
            role=role,
            status=AdminUser.STATUS_ACTIVE,
        )
        self.db.add(user)
        self.db.flush()
        publish_audit(
            self.db,
            admin=admin,
            action="staff.create",
            target_type="admin_user",
            target_id=str(user.id),
            detail={"username": username, "role": role},
            reason=f"创建员工账号 {username}",
        )
        self.db.commit()
        return user

    def _get(self, admin: AdminUser, user_id: int) -> AdminUser:
        user = (
            self.db.query(AdminUser)
            .filter(AdminUser.id == user_id, AdminUser.is_deleted == 0)
            .first()
        )
        if not user:
            raise NotFoundError("员工不存在")
        return user

    def update(
        self, admin: AdminUser, user_id: int, display_name: str | None, role: str | None
    ) -> AdminUser:
        user = self._get(admin, user_id)
        if user.id == admin.id and role and role != user.role:
            raise ValidationError("不能修改自己的角色（防自杀锁）")
        if role is not None:
            if role not in (AdminUser.ROLE_SUPER_ADMIN, AdminUser.ROLE_STAFF):
                raise ValidationError("角色必须是 superadmin 或 staff")
            user.role = role
        if display_name is not None:
            user.display_name = display_name
        publish_audit(
            self.db,
            admin=admin,
            action="staff.update",
            target_type="admin_user",
            target_id=str(user.id),
            detail={"display_name": display_name, "role": role},
            reason=f"更新员工账号（id={user.id}）",
        )
        self.db.commit()
        return user

    def set_status(self, admin: AdminUser, user_id: int, status: int) -> AdminUser:
        user = self._get(admin, user_id)
        if user.id == admin.id:
            raise ValidationError("不能禁用自己（防自杀锁）")
        user.status = AdminUser.STATUS_ACTIVE if status else AdminUser.STATUS_DISABLED
        publish_audit(
            self.db,
            admin=admin,
            action="staff.status",
            target_type="admin_user",
            target_id=str(user.id),
            detail={"status": status},
            reason=f"{'启用' if status else '禁用'}员工账号（{user.username}）",
        )
        self.db.commit()
        return user

    def reset_password(self, admin: AdminUser, user_id: int, new_password: str) -> None:
        user = self._get(admin, user_id)
        user.password_hash = hash_password(new_password)
        publish_audit(
            self.db,
            admin=admin,
            action="staff.reset_password",
            target_type="admin_user",
            target_id=str(user.id),
            detail={"username": user.username},
            reason=f"重置员工密码（{user.username}）",
        )
        self.db.commit()
