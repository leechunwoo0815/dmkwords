# backend/domain/catalog/import_service.py — Excel 批量导入（R-316 / V1.1 §4.3）
"""逐行校验：错误行报"第 N 行：原因"，不影响其他行；同 ISBN = 增副本不重建书目。"""

from __future__ import annotations

import io
import re

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from backend.domain.catalog.audit_events import publish_audit
from backend.domain.catalog.models import Book, BookCopy
from backend.domain.catalog.repository import BookCopyRepository, BookRepository

ISBN_RE = re.compile(r"^\d{9}[\dXx]$|^\d{13}$")

HEADER = ["ISBN", "书名", "作者", "AR值", "总词数", "主题", "年级", "副本数量"]


def _cell_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _max_copy_seq(copy_repo: BookCopyRepository, book: Book) -> int:
    prefix = f"C{(book.isbn or book.internal_code or 'BK')[-8:]}-"
    codes = [
        c
        for (c,) in copy_repo.db.query(BookCopy.copy_code).filter(BookCopy.book_id == book.id).all()
    ]
    best = 0
    for code in codes:
        if code and code.startswith(prefix):
            try:
                best = max(best, int(code[len(prefix) :]))
            except ValueError:
                continue
    return best


def import_books(db: Session, admin, file_bytes: bytes) -> dict:
    """返回 {total, success, failed, errors: ["第N行: 原因"]}。"""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    book_repo = BookRepository(db)
    copy_repo = BookCopyRepository(db)

    errors: list[str] = []
    success = 0
    row_num = 0
    created_books: list[dict] = []
    added_copies: list[dict] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_num += 1
        if not any(_cell_str(c) for c in row):
            continue  # 跳过空行
        isbn, title, author, ar, word_count, topic, grade, copies = (
            _cell_str(row[0]) if len(row) > 0 else "",
            _cell_str(row[1]) if len(row) > 1 else "",
            _cell_str(row[2]) if len(row) > 2 else "",
            _cell_str(row[3]) if len(row) > 3 else "",
            _cell_str(row[4]) if len(row) > 4 else "",
            _cell_str(row[5]) if len(row) > 5 else "",
            _cell_str(row[6]) if len(row) > 6 else "",
            _cell_str(row[7]) if len(row) > 7 else "1",
        )

        # ---- 行级校验 ----
        if not title:
            errors.append(f"第{row_num}行: 书名不能为空")
            continue
        if isbn and not ISBN_RE.match(isbn):
            errors.append(f"第{row_num}行: ISBN 格式不正确（{isbn}）")
            continue
        try:
            wc = int(float(word_count)) if word_count else 0
            if wc < 0:
                raise ValueError
        except ValueError:
            errors.append(f"第{row_num}行: 总词数必须是正整数")
            continue
        try:
            copy_n = int(float(copies)) if copies else 1
            if copy_n < 1 or copy_n > 99:
                raise ValueError
        except ValueError:
            errors.append(f"第{row_num}行: 副本数量须在 1-99 之间")
            continue

        if isbn:
            existing = book_repo.get_by_isbn(isbn)
            if existing:
                # 同 ISBN = 增副本（V1.1 §4.1）
                seq = _max_copy_seq(copy_repo, existing) + 1
                for i in range(copy_n):
                    copy_repo.create(
                        BookCopy(
                            book_id=existing.id,
                            copy_code=copy_repo.next_copy_code(existing, seq + i),
                        )
                    )
                added_copies.append({"isbn": isbn, "added": copy_n})
                success += 1
                continue

        book = Book(
            isbn=isbn or None,
            title=title,
            author=author,
            word_count=wc,
            ar_level=ar or None,
            topic=topic,
            grade=grade,
            status=Book.STATUS_ON,
        )
        book_repo.create(book)
        if not isbn:
            book.internal_code = f"LOCAL-{book.id:06d}"
            book_repo.update(book)
        seq = _max_copy_seq(copy_repo, book) + 1
        for i in range(copy_n):
            copy_repo.create(
                BookCopy(book_id=book.id, copy_code=copy_repo.next_copy_code(book, seq + i))
            )
        created_books.append({"title": title, "isbn": isbn})
        success += 1

    if created_books or added_copies:
        publish_audit(
            db,
            admin=admin,
            action="book.import",
            target_type="book",
            target_id="batch",
            detail={
                "created": len(created_books),
                "copies_added": len(added_copies),
                "failed": len(errors),
            },
            reason="Excel 批量导入",
        )
    db.commit()
    wb.close()
    return {
        "total_rows": row_num,
        "success_count": success,
        "failed_count": len(errors),
        "errors": errors[:50],
    }
