# tests/unit/test_wm11_dashboard_export.py — 数据看板补全 + Excel 导出（C18/D5）
"""断言锚点：PRD §11.2 / FEAT-069 / docs/04 WM11 步骤 6-7 / C18 审计导出。
- 看板新字段：维护/遗失/续费率/退会率/测验通过率/里程碑/待评估
- 导出：审计日志 / 数据看板 / 通知记录 三份 Excel 可下载可打开
"""

from datetime import datetime, timedelta

from fastapi.testclient import TestClient

from backend.database import get_session
from backend.domain.circulation.models import BorrowRecord
from tests.unit.helpers import force_book_on


def _h(client, username="admin"):
    r = client.post("/api/admin/login", json={"username": username, "password": "dmkwords123"})
    return {"Authorization": f"Bearer {r.json()['token']}"}


def _mk(client, h, phone="13800004001"):
    p = client.post(
        "/api/admin/members/parents", json={"name": "家长", "phone": phone}, headers=h
    ).json()
    c = client.post(
        f"/api/admin/members/parents/{p['id']}/children", json={"name": "看板孩"}, headers=h
    ).json()
    o = client.post(
        "/api/admin/orders", json={"child_id": c["id"], "order_type": "observation_fee"}, headers=h
    ).json()
    client.post(
        f"/api/admin/orders/{o['id']}/confirm-payment", json={"pay_method": "scan"}, headers=h
    )
    do = client.post(f"/api/admin/deposits/children/{c['id']}/orders", headers=h).json()
    client.post(
        f"/api/admin/orders/{do['order_id']}/confirm-payment",
        json={"pay_method": "scan"},
        headers=h,
    )
    book = client.post(
        "/api/admin/books",
        json={"isbn": "9780545582889", "title": "Dog Man", "word_count": 2500},
        headers=h,
    ).json()
    force_book_on(client, h, book["id"])
    return p, c, book


def test_dashboard_new_fields(client: TestClient):
    h = _h(client)
    p, c, book = _mk(client, h)
    # 借出 → 在借；再借一本直改 renew_used 造续费；逾期一本
    client.post(
        "/api/admin/circulation/borrow", json={"child_id": c["id"], "isbn": book["isbn"]}, headers=h
    )
    b2 = client.post(
        "/api/admin/books",
        json={"isbn": "9780747532699", "title": "Harry", "word_count": 1000},
        headers=h,
    ).json()
    force_book_on(client, h, b2["id"])
    r2 = client.post(
        "/api/admin/circulation/borrow", json={"child_id": c["id"], "isbn": b2["isbn"]}, headers=h
    )
    # 直改：借阅1 已还+续借（续费率）；借阅2 逾期（逾期数）
    with get_session() as db:
        rec1 = db.query(BorrowRecord).order_by(BorrowRecord.id.desc()).offset(1).first()
        rec1.status = BorrowRecord.STATUS_RETURNED
        rec1.returned_at = datetime.now()
        rec1.renew_used = 1
        rec2 = db.query(BorrowRecord).filter(BorrowRecord.id == r2.json()["id"]).first()
        rec2.due_at = datetime.now() - timedelta(days=1)
        db.commit()
    # 里程碑 + 测验通过率直改
    with get_session() as db:
        from backend.domain.growth.models import MilestoneAward, QuizAttempt

        db.add(MilestoneAward(child_id=c["id"], node_words=100000))
        db.add(
            QuizAttempt(
                child_id=c["id"],
                book_id=book["id"],
                score=4,
                total_questions=5,
                passed=1,
                snapshot="[]",
                submitted_at=datetime.now(),
            )
        )
        db.add(
            QuizAttempt(
                child_id=c["id"],
                book_id=b2["id"],
                score=2,
                total_questions=5,
                passed=0,
                snapshot="[]",
                submitted_at=datetime.now(),
            )
        )
        db.commit()

    resp = client.get("/api/admin/dashboard", headers=h)
    assert resp.status_code == 200
    d = resp.json()
    assert d["copy_maintenance"] == 0
    assert d["copy_lost"] == 0
    assert d["renew_rate"] == 100.0  # 唯一已还记录有续借 → 100%
    assert d["overdue_active"] >= 1
    assert d["quiz_pass_rate"] == 50.0
    assert d["milestone_count"] >= 1
    assert "pending_evaluation_count" in d


def test_export_audit_dashboard_notifications(client: TestClient):
    h = _h(client)
    p, c, book = _mk(client, h)
    client.post(
        "/api/admin/circulation/borrow", json={"child_id": c["id"], "isbn": book["isbn"]}, headers=h
    )
    # 审计日志导出（超管 audit.view）
    r = client.get("/api/admin/audit-logs/export", headers=h)
    assert r.status_code == 200
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert len(r.content) > 1000
    # 数据看板导出
    r2 = client.get("/api/admin/dashboard/export", headers=h)
    assert r2.status_code == 200
    assert len(r2.content) > 1000
    # 通知记录导出
    r3 = client.get("/api/admin/notifications/export", headers=h)
    assert r3.status_code == 200
    assert len(r3.content) > 1000


def test_dashboard_export_openable(client: TestClient):
    h = _h(client)
    p, c, book = _mk(client, h, phone="13800004002")
    r = client.get("/api/admin/dashboard/export", headers=h)
    assert r.status_code == 200
    from io import BytesIO

    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(r.content))
    ws = wb.active
    assert ws.title == "数据看板"
    labels = [row[0].value for row in ws.iter_rows()]
    assert "总藏书量" in labels
    assert "测验通过率(%)" in labels
