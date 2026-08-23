# tests/unit/test_p1_crud.py — P1 批次后端验收（C8 模板 / C12 员工管理 / C19 孩子资料）
"""真实 MySQL + TestClient；每项服务端行为断言，避免手工改库绕过。"""

from io import BytesIO

from fastapi.testclient import TestClient


def _h(client: TestClient, username: str = "admin") -> dict:
    r = client.post("/api/admin/login", json={"username": username, "password": "dmkwords123"})
    assert r.status_code == 200, r.text
    return {"Authorization": f"Bearer {r.json()['token']}"}


def test_import_template_download(client: TestClient):
    """C8：模板接口 200 + xlsx 可读（表头与 import_books 解析列序一致）。"""
    h = _h(client)
    r = client.get("/api/admin/books/import-template", headers=h)
    assert r.status_code == 200, r.text
    assert "spreadsheetml" in r.headers["content-type"]
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(r.content))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert headers[0] == "ISBN"
    assert headers[1] == "书名*"
    assert headers[7] == "副本数"


def test_staff_crud_permissions(client: TestClient):
    """C12：超管可建/改/禁/重置密码；staff 无权限（403）。"""
    h = _h(client)
    # staff 角色被拒绝
    hs = _h(client, "staff01")
    assert client.get("/api/admin/staff", headers=hs).status_code == 403
    # 超管创建
    r = client.post(
        "/api/admin/staff",
        json={
            "username": "staff02",
            "password": "dmkwords124",
            "display_name": "运营02",
            "role": "staff",
        },
        headers=h,
    )
    assert r.status_code == 200, r.text
    uid = r.json()["id"]
    # 新建员工可登录
    rl = client.post("/api/admin/login", json={"username": "staff02", "password": "dmkwords124"})
    assert rl.status_code == 200, rl.text
    # 重复用户名拒绝
    assert (
        client.post(
            "/api/admin/staff",
            json={
                "username": "staff02",
                "password": "dmkwords1244",
                "display_name": "重复",
                "role": "staff",
            },
            headers=h,
        ).status_code
        == 409
    )
    # 改名 + 禁用 + 密码重置
    assert (
        client.put(
            f"/api/admin/staff/{uid}", json={"display_name": "运营02改", "role": "staff"}, headers=h
        ).status_code
        == 200
    )
    assert (
        client.put(f"/api/admin/staff/{uid}/status", json={"status": 0}, headers=h).status_code
        == 200
    )
    # 禁用的员工登录被拒（403 账号已禁用）
    assert (
        client.post(
            "/api/admin/login", json={"username": "staff02", "password": "dmkwords124"}
        ).status_code
        == 403
    )
    assert (
        client.put(f"/api/admin/staff/{uid}/status", json={"status": 1}, headers=h).status_code
        == 200
    )
    assert (
        client.post(
            f"/api/admin/staff/{uid}/reset-password",
            json={"new_password": "dmkwords125"},
            headers=h,
        ).status_code
        == 200
    )
    assert (
        client.post(
            "/api/admin/login", json={"username": "staff02", "password": "dmkwords125"}
        ).status_code
        == 200
    )
    # 不能禁用自己
    me = client.get("/api/admin/me", headers=h).json()["user"]
    assert (
        client.put(f"/api/admin/staff/{me['id']}/status", json={"status": 0}, headers=h).status_code
        == 422
    )


def test_child_update_ar_only_up(client: TestClient):
    """C19：资料更新（english_name/grade/ar_level）；AR 只升不降；降级 422；审计留痕。"""
    h = _h(client)
    p = client.post(
        "/api/admin/members/parents", json={"name": "家长", "phone": "13800001401"}, headers=h
    ).json()
    c = client.post(
        f"/api/admin/members/parents/{p['id']}/children", json={"name": "AR孩"}, headers=h
    ).json()
    # 首次设置 + 年级
    r = client.put(
        f"/api/admin/members/children/{c['id']}",
        json={"english_name": "Tom", "grade": "G2", "ar_level": "3.5"},
        headers=h,
    )
    assert r.status_code == 200, r.text
    assert r.json()["english_name"] == "Tom"
    assert r.json()["ar_level"] == "3.5"
    # 上调允许
    r2 = client.put(f"/api/admin/members/children/{c['id']}", json={"ar_level": "4.0"}, headers=h)
    assert r2.status_code == 200
    assert r2.json()["ar_level"] == "4.0"
    # 下调拒绝
    r3 = client.put(f"/api/admin/members/children/{c['id']}", json={"ar_level": "3.0"}, headers=h)
    assert r3.status_code == 422
    assert "只升不降" in r3.json()["detail"]
    # 非法值拒绝
    assert (
        client.put(
            f"/api/admin/members/children/{c['id']}", json={"ar_level": "abc"}, headers=h
        ).status_code
        == 422
    )
    # 审计
    logs = client.get(
        "/api/admin/audit-logs", params={"action": "child.update_profile"}, headers=h
    ).json()
    assert logs["total"] >= 2
